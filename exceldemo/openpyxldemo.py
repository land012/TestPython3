# coding: utf-8
"""
# Created by xudazhou at 2020/2/22
"""
import openpyxl
import logging
import unittest


logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s %(levelname)s (%(lineno)d) - %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')


class OpenpyXlDeom(unittest.TestCase):

    @staticmethod
    def test1():
        """
        版本号
        行数 列数
        :return:
        """
        logging.info(openpyxl.__version__)  # 2.6.1
        filename = r"E:\TDDOWNLOAD\hadoop任务收集-feed架构.xlsx"
        wb1 = openpyxl.load_workbook(filename)

        logging.info(wb1.sheetnames)
        sheet1 = wb1.get_sheet_by_name("工作表1")
        logging.info(sheet1.max_row)  # 51
        logging.info(sheet1.max_column)  # 11

    @staticmethod
    def test2():
        """
        读一个 Cell
        :return:
        """
        filename = r"E:\TDDOWNLOAD\hadoop任务收集-feed架构.xlsx"
        wb1 = openpyxl.load_workbook(filename)

        logging.info(wb1.sheetnames)
        sheet1 = wb1.get_sheet_by_name("工作表1")

        cell1 = sheet1.cell(row=2, column=1)
        logging.info(cell1.value)  # wuzhonggang_shoubai_nonnews_uniq_bloom_$date
        logging.info(cell1.coordinate)  # A2

        cell1 = sheet1.cell(row=3, column=1)
        logging.info(cell1.value)  # None
        logging.info(cell1.coordinate)  # A3

    @staticmethod
    def test2_2():
        """
        合并 Cell
        :return:
        """
        filename = r"E:\TDDOWNLOAD\hadoop任务收集-feed架构.xlsx"
        wb1 = openpyxl.load_workbook(filename)
        sheet1 = wb1.get_sheet_by_name("工作表1")

        cell1 = sheet1.cell(row=2, column=1)
        logging.info(cell1.value)  # wuzhonggang_shoubai_nonnews_uniq_bloom_$date
        logging.info(type(cell1.parent))  # <class 'openpyxl.worksheet.worksheet.Worksheet'>
        logging.info(type(cell1))  # <class 'openpyxl.cell.cell.Cell'>

        cell2 = sheet1.cell(row=3, column=1)
        logging.info(cell2.value)  # None
        type2 = type(cell2)
        logging.info(type2)  # <class 'openpyxl.cell.cell.MergedCell'>
        logging.info(type2 == "openpyxl.cell.cell.MergedCell")  # False
        logging.info(type2 == openpyxl.cell.cell.MergedCell)  # True
        logging.info(isinstance(cell2, openpyxl.cell.cell.Cell))  # False
        logging.info(isinstance(cell2, openpyxl.cell.cell.MergedCell))  # True
        # logging.info(type2.tp_name)
        logging.info(type(type2))  # <class 'type'>

    @staticmethod
    def test3():
        """
        遍历指定区域
        :return:
        """
        filename = r"E:\TDDOWNLOAD\hadoop任务收集-feed架构.xlsx"
        wb1 = openpyxl.load_workbook(filename)

        logging.info(wb1.sheetnames)
        sheet1 = wb1.get_sheet_by_name("工作表1")

        range1 = sheet1['A1': 'F28']
        logging.info(type(range1))  # <class 'tuple'>
        logging.info(len(range1))  # 28
        logging.info(type(range1[0]))  # <class 'tuple'>
        logging.info(type(range1[0][0]))  # <class 'openpyxl.cell.cell.Cell'>

        for row in range1:
            for cell in row:
                print(cell.value, sep='', end='\t')
            print()

    @staticmethod
    def test4():
        """
        写 Cell
        :return:
        """
        logging.info(openpyxl.__version__)  # 2.6.1
        filename = r"E:\TDDOWNLOAD\hadoop任务收集-feed架构.xlsx"
        wb1 = openpyxl.load_workbook(filename)

        logging.info(wb1.sheetnames)
        sheet1 = wb1.get_sheet_by_name("工作表1")

        # 修改 excel
        cell2 = sheet1.cell(row=1, column=2)
        cell2.value = "job instance"
        wb1.save(filename=filename)

    @staticmethod
    def test4_2():
        """
        写 Excel
        :return:
        """
        filename = r"E:\TDDOWNLOAD\excel1.xlsx"
        wb1 = openpyxl.Workbook()

        sheet1 = wb1.create_sheet(title="sheet1")  # 创建了第2个 sheet
        sheet1["A1"] = "hello excel"
        wb1.save(filename=filename)

    @staticmethod
    def test4_3():
        """
        写 Excel
        :return:
        """
        filename = r"E:\TDDOWNLOAD\excel1.xlsx"
        wb1 = openpyxl.Workbook()

        sheet1 = wb1.active  # 获取第一个默认的 sheet
        sheet1.title = "title1"
        sheet1["A1"] = "hello excel"
        wb1.save(filename=filename)

    @staticmethod
    def test5():
        # (<class 'openpyxl.cell.cell.Cell'>, <class 'openpyxl.styles.styleable.StyleableObject'>, <class 'object'>)
        print(openpyxl.cell.cell.Cell.__mro__)

