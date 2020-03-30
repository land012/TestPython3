# coding: utf-8
"""
# Created by xudazhou at 2020/2/23
"""
import openpyxl
import logging


logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s %(levelname)s (%(lineno)d) - %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')


class Product:

    def __init__(self, name, category):
        self.name = name
        self.category = category

    def __hash__(self):
        return hash(self.name)

    def __eq__(self, other):
        return self.name == other.name and self.category == other.category

    def __str__(self):
        return self.name + "|" + self.category

    def __repr__(self):
        return self.__str__()


if __name__ == "__main__":
    wb2 = openpyxl.load_workbook(r"E:\TDDOWNLOAD\产品管理20200222A.xlsx")
    wb2_sheet1 = wb2["产品线映射关系"]

    dict1 = dict()  # 合同台账字典
    for row in wb2_sheet1.iter_rows(min_row=3):
        # if not row[1].value and row[2].value is None:
        #     logging.info(row[1].value)
        if row[1].value:
            p = Product(row[1].value, row[2].value)
            dict1[p] = row[0].value

    print(dict1)

    wb1 = openpyxl.load_workbook(r"E:\TDDOWNLOAD\2020合同台账表头-20200210A.xlsx")

    wb3 = openpyxl.Workbook()
    wb3_sheet1 = wb3.active
    wb3_sheet1.title = "sheet1"

    wb1_sheet2 = wb1["CRM合同台账"]
    last_a = ""
    last_b = ""
    last_f = ""
    last_g = ""
    last_n = ""
    last_p = ""
    last_ab = 0
    for row_num in range(5, 255):
        cell_a = wb1_sheet2.cell(row=row_num, column=1)
        cell_b = wb1_sheet2.cell(row=row_num, column=2)
        cell_f = wb1_sheet2.cell(row=row_num, column=6)
        cell_g = wb1_sheet2.cell(row=row_num, column=7)
        cell_n = wb1_sheet2.cell(row=row_num, column=14)
        cell_p = wb1_sheet2.cell(row=row_num, column=16)
        cell_u = wb1_sheet2.cell(row=row_num, column=21)
        cell_v = wb1_sheet2.cell(row=row_num, column=22)
        cell_x = wb1_sheet2.cell(row=row_num, column=24)
        cell_z = wb1_sheet2.cell(row=row_num, column=26)
        cell_ab = wb1_sheet2.cell(row=row_num, column=28)

        if isinstance(cell_a, openpyxl.cell.cell.MergedCell):
            pass
        else:
            wb3_sheet1.cell(row=row_num-4, column=1, value=cell_a.value)
            wb3_sheet1.cell(row=row_num - 4, column=2, value=cell_b.value)

            last_a = cell_a.value
            last_b = cell_b.value
            last_f = cell_f.value
            last_g = cell_g.value
            last_n = cell_n.value
            last_p = cell_p.value
            last_ab = cell_ab.value
